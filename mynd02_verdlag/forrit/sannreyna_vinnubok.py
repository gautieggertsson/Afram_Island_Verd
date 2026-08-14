#!/usr/bin/env python3
"""Sannreynir frystu gagnavinnubókina gegn frumgögnunum í ``hra/``.

Vinnubókin ``gogn/mynd02_verdlag_frosin.xlsx`` er fast inntak myndakóðans.
Þetta forrit les blaðið ``Inntak_myndar`` og ber hverja tölu saman við
frystu frumgögnin: verðlagsvísitölurnar A01 og E011 við Eurostat-svarið,
mældar miðgildistekjur 2024 við ``ilc_di03`` og framreiknaða íslenska
gildið við eigin endurreikning úr SILC 2020, þjóðhagsreikningum og
meðalgengi Seðlabankans. Ungverjaland er sérmerkt: talan sem endurgerir
birta reiti greiningarinnar er ályktuð af reitunum sjálfum, ekki sótt
gildi, og forritið ber hana aðeins saman við skjalfesta gildið 8.815 evrur.

Keyrsla:  python3 forrit/sannreyna_vinnubok.py
Forritið stöðvast á fyrstu ósamsvörun; annars prentar það yfirlit.
"""

from __future__ import annotations

import csv
import json
import math
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "hra"
BOK = ROOT / "gogn" / "mynd02_verdlag_frosin.xlsx"

HU_ALYKTAD = 8815.0


def jsonstat_values(path: Path, **fix: str) -> dict[str, float]:
    d = json.loads(path.read_text())
    dims = d["id"]
    sizes = d["size"]
    index = {dim: d["dimension"][dim]["category"]["index"] for dim in dims}
    strides: dict[str, int] = {}
    acc = 1
    for dim, size in zip(reversed(dims), reversed(sizes)):
        strides[dim] = acc
        acc *= size
    base = 0
    for dim, value in fix.items():
        base += index[dim][value] * strides[dim]
    out = {}
    for geo, gi in index["geo"].items():
        flat = base + gi * strides["geo"]
        val = d["value"].get(str(flat))
        if val is not None:
            out[geo] = float(val)
    return out


def framreiknad_island() -> float:
    silc = jsonstat_values(RAW / "eurostat_ilc_di03_is_2020.json", unit="NAC")
    silc_2020 = silc["IS"]

    heild = {}
    for line in (RAW / "hagstofa_frett_vintage_2025-04-04_PrrvG.csv").read_text().splitlines():
        cells = line.split("\t")
        if cells[0].strip() == "Ráðstöfunartekjur (tekjur - gjöld)":
            heild = {"2020": cells[1], "2024": cells[5]}
    v2020 = float(heild["2020"].replace(",", "."))
    v2024 = float(heild["2024"].replace(",", "."))

    gengi = []
    for line in (RAW / "sedlabanki_eur_midgengi_2024.csv").read_text().splitlines():
        cells = line.split(";")
        if len(cells) >= 8 and cells[4] == "Evra":
            gengi.append(float(cells[7]))
    medalgengi = sum(gengi) / len(gengi)

    return silc_2020 * (v2024 / v2020) / medalgengi


def main() -> None:
    a01 = jsonstat_values(RAW / "eurostat_prc_ppp_ind_2024.json", ppp_cat="A01")
    e011 = jsonstat_values(RAW / "eurostat_prc_ppp_ind_2024.json", ppp_cat="E011")
    tekjur_2024 = jsonstat_values(RAW / "eurostat_ilc_di03_2024_eur.json")

    wb = openpyxl.load_workbook(BOK, data_only=True)
    ws = wb["Inntak_myndar"]
    haus = [c.value for c in ws[4]]
    col = {name: i for i, name in enumerate(haus)}

    skodad = 0
    for row in ws.iter_rows(min_row=5, max_row=33, values_only=True):
        geo = row[col["geo"]]
        if geo is None:
            continue
        hlutverk = row[col["hlutverk"]]
        vA = row[col["pli_A01"]]
        vE = row[col["pli_E011"]]
        replik = row[col["midgildi_tekna_replik_eur"]]
        nuverandi = row[col["midgildi_tekna_nuverandi_eur"]]

        assert abs(vA - a01[geo]) < 1e-9, (geo, "A01", vA, a01[geo])
        assert abs(vE - e011[geo]) < 1e-9, (geo, "E011", vE, e011[geo])

        if nuverandi is not None and geo in tekjur_2024:
            assert abs(nuverandi - tekjur_2024[geo]) < 1e-6, (geo, "tekjur 2024")

        if hlutverk == "land" and geo not in ("IS", "HU"):
            assert abs(replik - tekjur_2024[geo]) < 1e-6, (geo, "replik")
        elif geo == "HU":
            assert abs(replik - HU_ALYKTAD) < 1e-6, ("HU", replik)
        elif geo == "IS":
            endurreiknad = framreiknad_island()
            assert abs(replik - endurreiknad) < 1e-6, ("IS", replik, endurreiknad)
        skodad += 1

    print(f"Vinnubókin stenst samanburð við frumgögnin: {skodad} raðir skoðaðar.")
    print("Ísland, framreiknað:", round(framreiknad_island(), 4), "evrur.")
    print("Ungverjaland: ályktað gildi", HU_ALYKTAD, "evrur (sjá SAMRAEMING.md).")


if __name__ == "__main__":
    main()
